import requests
from datetime import datetime
import schedule
import time
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import os

# API 키 (발급받은 API 키를 입력하세요)
api_key = '9ad4acf0420c46b0bba900de0cfbea94'  # 여기에 실제 API 키를 입력하세요
url = 'https://newsapi.org/v2/everything'  # 올바른 엔드포인트

# 주제별 색상 설정
topic_colors = {
    '경제': 'FFD700',  # 금색
    'IT': 'ADFF2F',    # 라이트 그린
    'AI': '1E90FF'     # 도저히 블루
}

def fetch_and_display_news():
    # 현재 날짜와 시간 가져오기
    current_time = datetime.now()
    formatted_time = current_time.strftime("%Y-%m-%d %H-%M-%S")

    print(f"크롤링 날짜 및 시간: {formatted_time}")
    print("API 요청을 시작합니다...")

    # HTTP GET 요청
    params = {'q': '경제 OR IT OR AI', 'language': 'ko', 'apiKey': api_key}
    response = requests.get(url, params=params)

    print("API 요청 완료, 상태 코드:", response.status_code)

    # 응답 상태 코드 확인
    if response.status_code == 200:
        print("응답이 성공적으로 수신되었습니다.")
        data = response.json()

        # 뉴스 기사를 담을 리스트
        articles = data.get('articles', [])

        # 주제별로 기사를 분류할 딕셔너리
        categorized_articles = {'경제': [], 'IT': [], 'AI': []}

        # 기사 중복 체크를 위한 링크 저장
        existing_links = set()

        # 기존 파일이 있는 경우, 기존 링크를 로드
        excel_file_name = "news_data.xlsx"
        if os.path.exists(excel_file_name):
            wb = load_workbook(excel_file_name)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, max_col=3, values_only=True):  # 2행부터 3열(A, B, C)까지
                existing_links.add(row[2])  # 링크만 저장

        # 기사 분류 및 중복 체크
        for article in articles:
            title = article['title'].replace('\n', ' ')  # 줄바꿈 제거
            link = article['url']
            topic = '경제' if '경제' in title else 'IT' if 'IT' in title else 'AI' if 'AI' in title else None

            if topic and link not in existing_links:  # 중복 체크
                categorized_articles[topic].append({
                    'title': title,
                    'link': link
                })

        # 각 주제별로 최대 5개 기사 저장
        all_articles = []
        for topic, articles in categorized_articles.items():
            if articles:
                for i, article in enumerate(articles[:5]):  # 최대 5개 저장
                    all_articles.append({
                        '주제': topic,
                        '제목': article['title'],
                        '링크': article['link']
                    })

        if all_articles:
            # 기존 파일이 있는지 확인
            if os.path.exists(excel_file_name):
                print(f"'{excel_file_name}' 파일이 존재합니다. 데이터를 추가합니다.")
                # 기존 파일 열기
                wb = load_workbook(excel_file_name)
                ws = wb.active

                # 새로운 데이터 추가
                for article in all_articles:
                    row = [article['주제'], article['제목'], article['링크']]
                    ws.append(row)

                    # 색상 설정
                    fill = PatternFill(start_color=topic_colors[article['주제']], end_color=topic_colors[article['주제']], fill_type='solid')
                    for col in range(1, 4):  # A, B, C 열에 색상 적용
                        ws.cell(row=ws.max_row, column=col).fill = fill

                # 생성된 날짜와 시간을 첫 번째 행에 추가
                ws.insert_rows(1)
                ws['A1'] = f"생성 날짜 및 시간: {current_time.strftime('%Y-%m-%d %H:%M:%S')}"
            else:
                print(f"'{excel_file_name}' 파일이 존재하지 않습니다. 새로 생성합니다.")
                # 새 엑셀 파일 생성
                df = pd.DataFrame(all_articles)
                df.to_excel(excel_file_name, index=False, engine='openpyxl')

                # 파일 열기
                wb = load_workbook(excel_file_name)
                ws = wb.active

                # 생성된 날짜와 시간을 첫 번째 행에 추가
                ws.insert_rows(1)
                ws['A1'] = f"생성 날짜 및 시간: {current_time.strftime('%Y-%m-%d %H:%M:%S')}"

                # 색상 설정
                for article in all_articles:
                    row = [article['주제'], article['제목'], article['링크']]
                    ws.append(row)

                    fill = PatternFill(start_color=topic_colors[article['주제']], end_color=topic_colors[article['주제']], fill_type='solid')
                    for col in range(1, 4):  # A, B, C 열에 색상 적용
                        ws.cell(row=ws.max_row, column=col).fill = fill

            # 파일 저장
            wb.save(excel_file_name)
            print(f"뉴스 기사를 '{excel_file_name}' 파일로 저장했습니다.")
            print("엑셀 파일 저장 완료.")
        else:
            print("저장할 뉴스 기사가 없습니다.")
    else:
        print("API 요청 실패, 상태 코드:", response.status_code)

# 12시와 00시에 뉴스 업데이트
schedule.every().day.at("00:00").do(fetch_and_display_news)
schedule.every().day.at("12:00").do(fetch_and_display_news)

# 프로그램 실행
fetch_and_display_news()  # 처음 실행하여 바로 보내기
print("초기 뉴스 크롤링 완료.")  # 초기 크롤링 완료 메시지
while True:
    schedule.run_pending()
    time.sleep(1)
